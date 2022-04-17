#excel_find
from openpyxl import load_workbook

dir= """D:\Python\study1excel.xlsx"""
wb = load_workbook("study1excel.xlsx")
ws = wb['Sheet1']

for row in ws.iter_rows(min_row=1) :
    if int(row[0].value )>5 :
        print(row[0].value) 