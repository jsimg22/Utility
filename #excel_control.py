#excel_control
from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()
ws.title = "testsheet1"
ws.sheet_properties.tabColor ="ff66ff"

ws2 = wb.create_sheet()
ws2.title = "testsheet1"
ws2.sheet_properties.tabColor ="11ffeb"

#ws = wb.active 
#ws.title = "ws.title"
wb.save("wb.test.xlsx")
wb.close()
