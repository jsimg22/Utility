#excel_loadfile
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

dir= """D:\Python\study1excel.xlsx"""
wb = load_workbook("study1excel.xlsx")
ws = wb['Sheet1']
ws['C1']="ADD"
ws.append(["Append1","Append2","Append3"])
wb.save(dir)


#for x in range(1,ws.max_row+1) :
#    for y in range(1,ws.max_column+1) :
#        print(ws.cell(row=x, column=y).value, end=" " )
#    print() 

